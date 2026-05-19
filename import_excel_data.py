#!/usr/bin/env python3
"""
Excel数据导入脚本
从 JiangWangNode.xlsx 读取三个僵尸网络的IP数据，进行富化并写入数据库
"""
import sys
import os
import asyncio
import logging
from datetime import datetime
from pathlib import Path

# 添加backend目录到路径
backend_dir = Path(__file__).parent / 'backend'
sys.path.insert(0, str(backend_dir))

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    print("缺少依赖包，正在安装...")
    os.system("pip3 install pandas openpyxl -q")
    import pandas as pd
    from openpyxl import load_workbook

from log_processor.enricher import IPEnricher
from log_processor.db_writer import BotnetDBWriter
from config import DB_CONFIG

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ExcelDataImporter:
    """Excel数据导入器"""
    
    def __init__(self, excel_file: str):
        """
        初始化导入器
        
        Args:
            excel_file: Excel文件路径
        """
        self.excel_file = excel_file
        self.enricher = IPEnricher(
            cache_size=20000,
            cache_ttl=86400,
            max_concurrent=50
        )
        
        # 工作表到僵尸网络类型的映射（全部小写）
        self.sheet_to_botnet = {
            'Autoupdate': 'autoupdate',
            'Asruex': 'asruex',
            'Mozi': 'mozi'
        }
        
    async def read_excel_sheet(self, sheet_name: str) -> list:
        """
        读取Excel工作表中的IP数据
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            IP地址列表
        """
        try:
            logger.info(f"正在读取工作表: {sheet_name}")
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            # 查找包含IP的列
            ip_column = None
            for col in df.columns:
                col_lower = str(col).lower()
                if 'ip' in col_lower or col_lower == 'ip':
                    ip_column = col
                    break
            
            if ip_column is None:
                # 如果没有找到IP列，尝试使用第一列
                ip_column = df.columns[0]
                logger.warning(f"未找到IP列，使用第一列: {ip_column}")
            
            # 提取IP地址并去重
            ips = df[ip_column].dropna().astype(str).unique().tolist()
            
            # 过滤无效IP
            valid_ips = []
            for ip in ips:
                ip = ip.strip()
                # 简单验证IP格式
                parts = ip.split('.')
                if len(parts) == 4 and all(p.isdigit() and 0 <= int(p) <= 255 for p in parts):
                    valid_ips.append(ip)
                else:
                    logger.debug(f"跳过无效IP: {ip}")
            
            logger.info(f"工作表 {sheet_name}: 找到 {len(valid_ips)} 个有效IP")
            return valid_ips
            
        except Exception as e:
            logger.error(f"读取工作表 {sheet_name} 失败: {e}")
            return []
    
    async def import_botnet_data(self, sheet_name: str, botnet_type: str):
        """
        导入单个僵尸网络的数据
        
        Args:
            sheet_name: Excel工作表名称
            botnet_type: 僵尸网络类型（小写）
        """
        try:
            logger.info(f"\n{'='*60}")
            logger.info(f"开始导入 {botnet_type} (工作表: {sheet_name})")
            logger.info(f"{'='*60}")
            
            # 读取IP列表
            ips = await self.read_excel_sheet(sheet_name)
            if not ips:
                logger.warning(f"工作表 {sheet_name} 中没有有效IP，跳过")
                return
            
            # 创建数据库写入器
            writer = BotnetDBWriter(botnet_type, DB_CONFIG, batch_size=500)
            
            # 当前时间作为所有记录的时间戳
            current_time = datetime.now()
            
            logger.info(f"开始处理 {len(ips)} 个IP...")
            
            # 批量处理IP
            batch_size = 100
            processed_count = 0
            error_count = 0
            
            for i in range(0, len(ips), batch_size):
                batch_ips = ips[i:i + batch_size]
                
                # 并发富化IP信息
                tasks = [self.enricher.enrich(ip) for ip in batch_ips]
                ip_infos = await asyncio.gather(*tasks, return_exceptions=True)
                
                # 写入数据库
                for ip, ip_info in zip(batch_ips, ip_infos):
                    try:
                        # 如果富化失败，使用默认信息
                        if isinstance(ip_info, Exception):
                            logger.warning(f"IP富化失败 {ip}: {ip_info}")
                            ip_info = {
                                'ip': ip,
                                'country': '未知',
                                'province': '',
                                'city': '',
                                'isp': '',
                                'asn': '',
                                'longitude': 0.0,
                                'latitude': 0.0,
                                'continent': '',
                                'is_china': False
                            }
                        
                        # 构造日志数据格式
                        log_data = {
                            'ip': ip,
                            'timestamp': current_time,
                            'event_type': 'excel_import',
                            'source': 'excel_importer',
                            'date': current_time.strftime('%Y-%m-%d'),
                            'botnet_type': botnet_type
                        }
                        
                        # 写入数据库
                        writer.add_node(log_data, ip_info)
                        processed_count += 1
                        
                    except Exception as e:
                        logger.error(f"处理IP {ip} 失败: {e}")
                        error_count += 1
                
                # 显示进度
                progress = min(i + batch_size, len(ips))
                logger.info(f"进度: {progress}/{len(ips)} ({progress*100//len(ips)}%)")
            
            # 强制刷新到数据库
            logger.info("正在写入数据库...")
            await writer.flush(force=True)
            
            # 显示统计信息
            stats = writer.get_stats()
            enricher_stats = self.enricher.get_stats()
            
            logger.info(f"\n{'='*60}")
            logger.info(f"{botnet_type} 导入完成")
            logger.info(f"{'='*60}")
            logger.info(f"总IP数: {len(ips)}")
            logger.info(f"成功处理: {processed_count}")
            logger.info(f"处理失败: {error_count}")
            logger.info(f"写入节点数: {stats['total_written']}")
            logger.info(f"去重数量: {stats['duplicate_count']}")
            logger.info(f"去重率: {stats['duplicate_rate']}")
            logger.info(f"IP富化缓存命中率: {enricher_stats['total_cache_hit_rate']}")
            logger.info(f"{'='*60}\n")
            
        except Exception as e:
            logger.error(f"导入 {botnet_type} 数据时发生错误: {e}", exc_info=True)
    
    async def import_all(self):
        """导入所有工作表的数据"""
        try:
            logger.info(f"开始导入Excel文件: {self.excel_file}")
            
            # 检查文件是否存在
            if not os.path.exists(self.excel_file):
                logger.error(f"文件不存在: {self.excel_file}")
                return
            
            # 读取所有工作表名称
            wb = load_workbook(self.excel_file, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            logger.info(f"发现工作表: {sheet_names}")
            
            # 处理每个工作表
            for sheet_name in sheet_names:
                if sheet_name in self.sheet_to_botnet:
                    botnet_type = self.sheet_to_botnet[sheet_name]
                    await self.import_botnet_data(sheet_name, botnet_type)
                else:
                    logger.warning(f"跳过未知工作表: {sheet_name}")
            
            logger.info("\n所有数据导入完成！")
            
        except Exception as e:
            logger.error(f"导入过程中发生错误: {e}", exc_info=True)


async def main():
    """主函数"""
    # Excel文件路径
    excel_file = os.path.join(os.path.dirname(__file__), 'JiangWangNode.xlsx')
    
    # 创建导入器
    importer = ExcelDataImporter(excel_file)
    
    # 执行导入
    await importer.import_all()


if __name__ == '__main__':
    # 运行导入任务
    asyncio.run(main())
